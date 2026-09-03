# -*- coding: utf-8 -*-
"""売上目標 再設定（税抜）: 広告費×ROAS＋自然売上 → ストア別・ブランド別の月次目標。数式で構築。"""
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as L

F = "Arial"; YEN = "¥#,##0;(¥#,##0);-"; PCT = "0.0%"; X = "0.0"
blue = Font(name=F, size=10, color="0000FF"); black = Font(name=F, size=10)
bold = Font(name=F, size=10, bold=True); green = Font(name=F, size=10, color="008000")
wb_ = Font(name=F, size=10, bold=True, color="FFFFFF"); small = Font(name=F, size=8, italic=True)
yellow = PatternFill("solid", fgColor="FFFF00"); hdr = PatternFill("solid", fgColor="1F3864")
sub = PatternFill("solid", fgColor="D9E1F2"); grey = PatternFill("solid", fgColor="EDEDED")
thin = Side(style="thin", color="BFBFBF"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

MONTHS = ["2026-09", "2026-10", "2026-11", "2026-12", "2027-01", "2027-02", "2027-03", "2027-04"]
COMPANY = [15_000_000, 29_000_000, 30_000_000, 36_000_000, 26_000_000, 41_000_000, 15_500_000, 16_500_000]
PREV = [9_504_761, 16_850_103, 28_306_061, 27_709_798, 22_725_973, 63_694_086, 14_646_068, 24_663_324]
AD_TOTAL = [750_000, 1_230_119, 1_271_123, 1_494_135, 1_148_111, 1_562_127, 993_091, 902_089]
EVENT = [100_000, 0, 0, 100_000, 0, 250_000, 50_000, 0]
GROWTH = [1.00, 1.10, 1.00, 1.00, 1.00, 0.75, 1.00, 0.90]  # 10月=3ストア稼働で+10%、2月=前年突出のため0.75、4月=前年高水準のため0.9
ACT = [("2026-05", 10_878_769, 16_500_000, 12_569_387), ("2026-06", 11_689_179, 16_500_000, 10_282_993),
       ("2026-07", 31_677_600, 17_000_000, 11_017_920), ("2026-08", 26_669_593, 21_000_000, 21_492_906)]
# ライン: (ストア, ブランド/枠, 9月配分比率(FULLMARKS通常予算内), 10月〜比率, ストア枠, ROAS税込, 自然シェア9月, 自然シェア10月〜, 備考)
LINES = [
    ("FULLMARKS STORE", "HOUDINI",           0.27, 0.20, "FM", 14.0, 0.55, 0.22, "8月ROAS17.3。HOUDINI STORE稼働後は自然売上の一部が移管"),
    ("FULLMARKS STORE", "NORRØNA",           0.07, 0.06, "FM", 4.0,  0.12, 0.08, "8月ROAS1.1→文言/シリーズ別に刷新。NORRONA STOREへ移管"),
    ("FULLMARKS STORE", "POC",               0.09, 0.12, "FM", 5.0,  0.10, 0.10, "8月ROAS3.9。専用ストア無しのため強化"),
    ("FULLMARKS STORE", "ACLIMA",            0.05, 0.06, "FM", 4.0,  0.06, 0.06, "MetaのCLKはROAS6"),
    ("FULLMARKS STORE", "HESTRA",            0.04, 0.08, "FM", 5.0,  0.07, 0.08, "未広告→指名検索を新設。10〜1月が本番"),
    ("FULLMARKS STORE", "その他(KANG/PLUS ONE WORKS/POW/SR)", 0.01, 0.02, "FM", 3.0, 0.10, 0.08, "テスト枠"),
    ("FULLMARKS STORE", "共通: 店舗指名検索", 0.17, 0.18, "FM", 14.0, 0.0, 0.0, "8月ROAS15.8。ブランド横断のため自然売上は配分しない"),
    ("FULLMARKS STORE", "共通: 商品連動(ショッピング/カタログ)", 0.30, 0.28, "FM", 8.0, 0.0, 0.0, "8月ROAS7.4。アウトレット除外後の水準で要観察"),
    ("FULLMARKS STORE", "共通: イベント予備費(SW/年末/セール)", None, None, "EV", 5.0, 0.0, 0.0, "EC企画の告知用。月次表の『イベント』列"),
    ("HOUDINI STORE",   "HOUDINI",           None, None, "ST", 8.0,  0.0, 0.25, "10月稼働。FULLMARKSのHOUDINI売上(71%)の一部が移管"),
    ("NORRONA STORE",   "NORRØNA",           None, None, "ST", 5.0,  0.0, 0.08, "10月稼働"),
    ("PU STORE",        "PLUS ONE WORKS",    None, None, "ST", 3.0,  0.0, 0.05, "10月稼働。小規模"),
]
STORE_SHARE = 0.15  # 10月以降、HOUDINI/NORRONA/PU 各15%（FULLMARKS 55%）

wb = Workbook()
# ============ 前提 ============
ws = wb.active; ws.title = "前提"
ws["A1"] = "売上目標 再設定 FY2026（2026-05〜2027-04）― 前提・入力値（すべて税抜）"; ws["A1"].font = Font(name=F, size=13, bold=True)
ws["A2"] = "青字＝入力値（変更可）、黄色＝特に見直しが必要な仮置き。黒字＝数式。他シートはこのシートを参照。"; ws["A2"].font = small
rows = [("税込→税抜 係数（媒体計測ROASは税込）", 1.1, "0.00", "媒体の購入金額は税込。÷1.1で税抜換算"),
        ("前年売上のうち広告経由の比率", 0.15, PCT, "前年は計測破損のため推定。8月実績の広告売り比率16%を参考に仮置き"),
        ("会社WEB予算（年間・税抜）", 280_000_000, YEN, "会社WEB年度予算ファイル"),
        ("ストレッチ目標（年間・税抜）", 320_000_000, YEN, "ユーザー指定"),
        ("3ストアの広告予算シェア（10月〜、各ストア）", STORE_SHARE, PCT, "HOUDINI/NORRONA/PU 各15%、FULLMARKS 55%")]
for i, (k, v, fmt, note) in enumerate(rows, 4):
    ws.cell(row=i, column=1, value=k).font = black
    c = ws.cell(row=i, column=2, value=v); c.font = blue; c.number_format = fmt
    if i in (5,): c.fill = yellow
    ws.cell(row=i, column=3, value=note).font = small
K_TAX, K_ADSHARE, K_BUDGET, K_STRETCH, K_STORE = "前提!$B$4", "前提!$B$5", "前提!$B$6", "前提!$B$7", "前提!$B$8"

# 実績 5〜8月
ws["A11"] = "実績（5〜8月・税抜）"; ws["A11"].font = bold
for c, h in enumerate(["月", "売上実績", "会社予算", "前年同月"], 1):
    cell = ws.cell(row=12, column=c, value=h); cell.font = wb_; cell.fill = hdr; cell.alignment = center
for i, (m, a, b, p) in enumerate(ACT, 13):
    ws.cell(row=i, column=1, value=m).font = black
    for c, v in enumerate((a, b, p), 2):
        cell = ws.cell(row=i, column=c, value=v); cell.font = blue; cell.number_format = YEN
ws["A17"] = "5〜8月 計"; ws["A17"].font = bold
for c in (2, 3, 4):
    cell = ws.cell(row=17, column=c, value=f"=SUM({L(c)}13:{L(c)}16)"); cell.font = bold; cell.number_format = YEN
ws["E13"] = "出典: 会社WEB年度予算ファイル（税抜）。予算比は5〜8月計で114%"; ws["E13"].font = small
YTD_ACT, YTD_BUD, YTD_PREV = "前提!$B$17", "前提!$C$17", "前提!$D$17"

# 月別入力 9〜4月
ws["A20"] = "月別入力（9〜4月）"; ws["A20"].font = bold
mh = ["月", "会社予算", "前年同月", "自然売上 成長率", "広告費合計(通常+イベント)", "うちイベント予備費"]
for c, h in enumerate(mh, 1):
    cell = ws.cell(row=21, column=c, value=h); cell.font = wb_; cell.fill = hdr; cell.alignment = center
M0 = 22
for i, m in enumerate(MONTHS):
    r = M0 + i
    ws.cell(row=r, column=1, value=m).font = black
    for c, v, fmt in ((2, COMPANY[i], YEN), (3, PREV[i], YEN), (4, GROWTH[i], "0.00"), (5, AD_TOTAL[i], YEN), (6, EVENT[i], YEN)):
        cell = ws.cell(row=r, column=c, value=v); cell.font = blue; cell.number_format = fmt
    ws.cell(row=r, column=4).fill = yellow
ws.cell(row=M0 + 8, column=1, value="計").font = bold
for c in (2, 3, 5, 6):
    cell = ws.cell(row=M0 + 8, column=c, value=f"=SUM({L(c)}{M0}:{L(c)}{M0+7})"); cell.font = bold; cell.number_format = YEN
ws["G22"] = ("成長率＝前年同月の自然売上に対する伸び（1.00＝前年並み）。広告費は「広告予算計画_金額確定版」の月額。"
             "2月は前年6,369万が突出値のため、成長率を下げる判断もあり得る。"); ws["G22"].font = small

# ライン定義
ws["A33"] = "ライン定義（広告配分比率・ROAS・自然売上シェア）"; ws["A33"].font = bold
lh = ["ストア", "ブランド/枠", "9月 配分比率", "10月〜 配分比率", "枠種別", "想定ROAS(税込計測)", "自然売上シェア 9月", "自然売上シェア 10月〜", "備考"]
for c, h in enumerate(lh, 1):
    cell = ws.cell(row=34, column=c, value=h); cell.font = wb_; cell.fill = hdr; cell.alignment = center
L0 = 35
for i, (store, brand, r9, r10, kind, roas, s9, s10, note) in enumerate(LINES):
    r = L0 + i
    ws.cell(row=r, column=1, value=store).font = black
    ws.cell(row=r, column=2, value=brand).font = black
    for c, v, fmt in ((3, r9, PCT), (4, r10, PCT), (6, roas, X), (7, s9, PCT), (8, s10, PCT)):
        cell = ws.cell(row=r, column=c, value=v); cell.font = blue; cell.number_format = fmt
        if c in (6, 8): cell.fill = yellow
    ws.cell(row=r, column=5, value=kind).font = black
    ws.cell(row=r, column=9, value=note).font = small
LR = L0 + len(LINES)
ws.cell(row=LR, column=1, value="チェック（比率・シェアの合計）").font = bold
for c in (3, 4, 7, 8):
    cell = ws.cell(row=LR, column=c, value=f"=SUM({L(c)}{L0}:{L(c)}{LR-1})"); cell.font = bold; cell.number_format = PCT
ws.cell(row=LR, column=9, value="配分比率は各100%、自然売上シェアは各100%になること").font = small
ws["A48"] = ("想定ROASの根拠: 8月実績（Google 16.1 / Meta 0.5、ブランド別は FULLMARKS内訳シート参照）を、"
             "計測復旧と文言刷新後の水準として保守的に置いた。新ストアは立ち上げ期のため低め。"); ws["A48"].font = small
for col, w in zip("ABCDEFGHI", (34, 34, 13, 14, 9, 16, 15, 16, 60)):
    ws.column_dimensions[col].width = w

# ============ ストア別ブランド別 ============
ws2 = wb.create_sheet("ストア別ブランド別")
ws2["A1"] = "ストア別・ブランド別 月次目標（税抜）＝ 広告費 × ROAS ÷ 1.1 ＋ 自然売上"; ws2["A1"].font = Font(name=F, size=13, bold=True)
ws2["A2"] = "各月4列: 広告費 / 広告売上 / 自然売上 / 合計。全て「前提」シートの入力値からの数式。"; ws2["A2"].font = small
# header rows 3-4
ws2.cell(row=4, column=1, value="ストア").font = wb_; ws2.cell(row=4, column=2, value="ブランド/枠").font = wb_
for c in (1, 2): ws2.cell(row=4, column=c).fill = hdr; ws2.cell(row=3, column=c).fill = hdr
subcols = ["広告費", "広告売上", "自然売上", "合計"]
for i, m in enumerate(MONTHS + ["年間(9〜4月)"]):
    c0 = 3 + i * 4
    ws2.merge_cells(start_row=3, start_column=c0, end_row=3, end_column=c0 + 3)
    cell = ws2.cell(row=3, column=c0, value=m); cell.font = wb_; cell.fill = hdr; cell.alignment = center
    for j, s in enumerate(subcols):
        cell = ws2.cell(row=4, column=c0 + j, value=s); cell.font = wb_; cell.fill = hdr; cell.alignment = center
R0 = 5
n = len(LINES)
for i, (store, brand, r9, r10, kind, roas, s9, s10, note) in enumerate(LINES):
    r = R0 + i; pr = L0 + i  # 前提シートの行
    ws2.cell(row=r, column=1, value=store).font = black
    ws2.cell(row=r, column=2, value=brand).font = black
    for mi, m in enumerate(MONTHS):
        c0 = 3 + mi * 4; mrow = M0 + mi
        tot = f"前提!$E${mrow}"; ev = f"前提!$F${mrow}"; normal = f"({tot}-{ev})"
        is_sep = "TRUE" if mi == 0 else "FALSE"
        if kind == "FM":
            ratio = f"IF({is_sep},前提!$C${pr},前提!$D${pr})"
            store_share = f"IF({is_sep},1,1-3*{K_STORE})"
            cost = f"={normal}*{store_share}*{ratio}"
        elif kind == "EV":
            cost = f"={ev}"
        else:
            cost = f"=IF({is_sep},0,{normal}*{K_STORE})"
        cA = ws2.cell(row=r, column=c0, value=cost)
        cB = ws2.cell(row=r, column=c0 + 1, value=f"={L(c0)}{r}*前提!$F${pr}/{K_TAX}")
        share = f"IF({is_sep},前提!$G${pr},前提!$H${pr})"
        cC = ws2.cell(row=r, column=c0 + 2, value=f"=月別目標!$D${6+mi}*{share}")
        cD = ws2.cell(row=r, column=c0 + 3, value=f"={L(c0+1)}{r}+{L(c0+2)}{r}")
        for cell in (cA, cB, cC): cell.font = green; cell.number_format = YEN
        cD.font = bold; cD.number_format = YEN
    # 年間
    c0 = 3 + 8 * 4
    for j in range(4):
        refs = "+".join(f"{L(3+mi*4+j)}{r}" for mi in range(8))
        cell = ws2.cell(row=r, column=c0 + j, value=f"={refs}"); cell.font = bold if j == 3 else black; cell.number_format = YEN
RT = R0 + n
ws2.cell(row=RT, column=1, value="4ストア合計").font = bold
for c in range(3, 3 + 9 * 4):
    cell = ws2.cell(row=RT, column=c, value=f"=SUM({L(c)}{R0}:{L(c)}{RT-1})"); cell.font = bold; cell.number_format = YEN; cell.fill = sub
# ストア小計
stores = ["FULLMARKS STORE", "HOUDINI STORE", "NORRONA STORE", "PU STORE"]
ws2.cell(row=RT + 2, column=1, value="ストア別 小計").font = bold
for si, st in enumerate(stores):
    r = RT + 3 + si
    ws2.cell(row=r, column=1, value=st).font = black
    for c in range(3, 3 + 9 * 4):
        cell = ws2.cell(row=r, column=c, value=f'=SUMIF($A${R0}:$A${RT-1},$A{r},{L(c)}${R0}:{L(c)}${RT-1})')
        cell.font = black; cell.number_format = YEN
# ROAS行（年間）
r = RT + 8
ws2.cell(row=r, column=1, value="年間 想定ROAS（税抜ベース＝広告売上÷広告費）").font = bold
c0 = 3 + 8 * 4
cell = ws2.cell(row=r, column=c0 + 1, value=f"=IF({L(c0)}{RT}=0,0,{L(c0+1)}{RT}/{L(c0)}{RT})"); cell.font = bold; cell.number_format = X
ws2.column_dimensions["A"].width = 18; ws2.column_dimensions["B"].width = 34
for c in range(3, 3 + 9 * 4): ws2.column_dimensions[L(c)].width = 12
ws2.freeze_panes = "C5"
for rr in range(3, RT + 7):
    for c in range(1, 3 + 9 * 4):
        ws2.cell(row=rr, column=c).border = border
ADREV_COL = lambda mi: L(3 + mi * 4 + 1); COST_COL = lambda mi: L(3 + mi * 4)

# ============ 月別目標 ============
ws3 = wb.create_sheet("月別目標", 1)
ws3["A1"] = "月別 売上目標（税抜）― 広告費 × ROAS ＋ 自然売上 vs 会社予算2.8億"; ws3["A1"].font = Font(name=F, size=13, bold=True)
ws3["A2"] = ("自然売上＝前年同月 × (1−前年広告比率) × 成長率。広告売上＝各ラインの広告費×想定ROAS÷1.1の合計（ストア別ブランド別シート）。"
             "目標＝自然売上＋広告売上。"); ws3["A2"].font = small
h3 = ["月", "会社予算", "前年同月", "自然売上", "広告費", "広告売上", "売上目標", "予算差", "前年比", "広告費率", "累計 目標", "累計 会社予算", "累計 予算差", "備考"]
for c, h in enumerate(h3, 1):
    cell = ws3.cell(row=4, column=c, value=h); cell.font = wb_; cell.fill = hdr; cell.alignment = center
# 実績行（5〜8月まとめ）
ws3.cell(row=5, column=1, value="5〜8月 実績").font = bold
for c, f in ((2, f"={YTD_BUD}"), (3, f"={YTD_PREV}"), (7, f"={YTD_ACT}"), (8, f"=G5-B5"), (9, f"=G5/C5"), (11, "=G5"), (12, "=B5"), (13, "=K5-L5")):
    cell = ws3.cell(row=5, column=c, value=f); cell.font = green if c in (2, 3, 7) else black
    cell.number_format = PCT if c == 9 else YEN
ws3.cell(row=5, column=14, value="実績。広告売り/自然売りの内訳は計測破損のため省略").font = small
for c in range(1, 15): ws3.cell(row=5, column=c).fill = grey
notes = ["SW企画。前年950万は年間最弱月", "3ストア稼働・広告増額。前年1,685万", "繁忙期入り", "年末商戦（予備10万）",
         "冬物実需", "冬セール（予備25万）。前年6,369万は突出値", "端境期（予備5万）", "調整月"]
for mi, m in enumerate(MONTHS):
    r = 6 + mi; mrow = M0 + mi
    ws3.cell(row=r, column=1, value=m).font = black
    ws3.cell(row=r, column=2, value=f"=前提!B{mrow}").font = green
    ws3.cell(row=r, column=3, value=f"=前提!C{mrow}").font = green
    ws3.cell(row=r, column=4, value=f"=C{r}*(1-{K_ADSHARE})*前提!D{mrow}").font = black
    ws3.cell(row=r, column=5, value=f"=ストア別ブランド別!{COST_COL(mi)}{RT}").font = green
    ws3.cell(row=r, column=6, value=f"=ストア別ブランド別!{ADREV_COL(mi)}{RT}").font = green
    ws3.cell(row=r, column=7, value=f"=D{r}+F{r}").font = bold
    ws3.cell(row=r, column=8, value=f"=G{r}-B{r}").font = black
    ws3.cell(row=r, column=9, value=f"=IF(C{r}=0,0,G{r}/C{r})").font = black
    ws3.cell(row=r, column=10, value=f"=IF(G{r}=0,0,E{r}/G{r})").font = black
    ws3.cell(row=r, column=11, value=f"=K{r-1}+G{r}").font = black
    ws3.cell(row=r, column=12, value=f"=L{r-1}+B{r}").font = black
    ws3.cell(row=r, column=13, value=f"=K{r}-L{r}").font = black
    ws3.cell(row=r, column=14, value=notes[mi]).font = small
    for c in range(2, 14):
        ws3.cell(row=r, column=c).number_format = PCT if c in (9, 10) else YEN
RY = 14
ws3.cell(row=RY, column=1, value="年間 合計").font = bold
for c, f in ((2, "=B5+SUM(B6:B13)"), (3, "=C5+SUM(C6:C13)"), (4, "=SUM(D6:D13)"), (5, "=SUM(E6:E13)"), (6, "=SUM(F6:F13)"),
             (7, "=G5+SUM(G6:G13)"), (8, f"=G{RY}-B{RY}"), (9, f"=G{RY}/C{RY}"), (10, f"=E{RY}/G{RY}")):
    cell = ws3.cell(row=RY, column=c, value=f); cell.font = bold; cell.number_format = PCT if c in (9, 10) else YEN; cell.fill = sub
ws3.cell(row=RY + 2, column=1, value="年間目標 vs 会社予算2.8億").font = bold
ws3.cell(row=RY + 2, column=7, value=f"=G{RY}-{K_BUDGET}").number_format = YEN
ws3.cell(row=RY + 3, column=1, value="年間目標 vs ストレッチ3.2億").font = bold
ws3.cell(row=RY + 3, column=7, value=f"=G{RY}-{K_STRETCH}").number_format = YEN
ws3.cell(row=RY + 4, column=1, value="年間 広告売上ROAS（税抜）").font = bold
ws3.cell(row=RY + 4, column=7, value=f"=F{RY}/E{RY}").number_format = X
for c, w in zip("ABCDEFGHIJKLMN", (14, 14, 14, 14, 12, 14, 14, 13, 9, 9, 15, 15, 14, 40)):
    ws3.column_dimensions[c].width = w
for rr in range(4, RY + 1):
    for c in range(1, 15): ws3.cell(row=rr, column=c).border = border

# ============ シナリオ ============
ws4 = wb.create_sheet("シナリオ")
ws4["A1"] = "シナリオ比較（年間売上・税抜）"; ws4["A1"].font = Font(name=F, size=13, bold=True)
ws4["A2"] = "自然売上とROASの達成度を掛けて年間売上を試算。青字は変更可。"; ws4["A2"].font = small
for c, h in enumerate(["シナリオ", "自然売上 達成度", "ROAS 達成度", "年間売上", "vs 2.8億", "vs 3.2億", "9〜4月 必要な月平均"], 1):
    cell = ws4.cell(row=4, column=c, value=h); cell.font = wb_; cell.fill = hdr; cell.alignment = center
for i, (name, og, rk) in enumerate([("保守（前年割れ・ROAS7割）", 0.95, 0.7), ("標準（前提シートどおり）", 1.0, 1.0), ("強気（自然+5%・ROAS1.3倍）", 1.05, 1.3)]):
    r = 5 + i
    ws4.cell(row=r, column=1, value=name).font = black
    ws4.cell(row=r, column=2, value=og).font = blue; ws4.cell(row=r, column=2).number_format = "0.00"
    ws4.cell(row=r, column=3, value=rk).font = blue; ws4.cell(row=r, column=3).number_format = "0.00"
    ws4.cell(row=r, column=4, value=f"={YTD_ACT}+月別目標!$D${RY}*B{r}+月別目標!$F${RY}*C{r}").number_format = YEN
    ws4.cell(row=r, column=5, value=f"=D{r}-{K_BUDGET}").number_format = YEN
    ws4.cell(row=r, column=6, value=f"=D{r}-{K_STRETCH}").number_format = YEN
    ws4.cell(row=r, column=7, value=f"=(D{r}-{YTD_ACT})/8").number_format = YEN
for c, w in zip("ABCDEFG", (30, 14, 12, 16, 15, 15, 18)): ws4.column_dimensions[c].width = w
for rr in range(4, 8):
    for c in range(1, 8): ws4.cell(row=rr, column=c).border = border

import sys
out = sys.argv[1] if len(sys.argv) > 1 else "reports/売上目標_再設定_2026-09.xlsx"
wb.calculation.fullCalcOnLoad = True
wb.save(out); print("saved", out)
